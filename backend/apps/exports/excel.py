import io
from decimal import Decimal
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from apps.operations.models import Operation, SoldeMois

MOIS_FR = [
    '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
]

# Couleurs
BLUE_BG = 'DDEEFF'
RED_BG = 'FFDDDD'
HEADER_BG = '2F4F8F'
TOTAL_BG = 'E8E8E8'
BORDER_COLOR = 'AAAAAA'


def thin_border():
    side = Side(style='thin', color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def generate_excel(mois: int, annee: int) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{MOIS_FR[mois]} {annee}'

    # --- Titre ---
    ws.merge_cells('A1:G1')
    titre = ws['A1']
    titre.value = f'JOURNAL DE CAISSE — {MOIS_FR[mois].upper()} {annee}'
    titre.font = Font(bold=True, size=14, color='FFFFFF')
    titre.fill = PatternFill('solid', fgColor=HEADER_BG)
    titre.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Organisation ---
    ws.merge_cells('A2:G2')
    org = ws['A2']
    org.value = 'EPA_OUAGA'
    org.font = Font(bold=True, size=11)
    org.alignment = Alignment(horizontal='center')

    # --- Solde reporté ---
    try:
        solde_obj = SoldeMois.objects.get(mois=mois, annee=annee)
        solde_reporte = solde_obj.solde_reporte
        total_entrees = solde_obj.total_entrees
        total_depenses = solde_obj.total_depenses
        solde_final = solde_obj.solde_final
    except SoldeMois.DoesNotExist:
        solde_reporte = Decimal('0.00')
        total_entrees = Decimal('0.00')
        total_depenses = Decimal('0.00')
        solde_final = Decimal('0.00')

    ws['A3'] = 'Solde reporté :'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = float(solde_reporte)
    ws['B3'].number_format = '#,##0.00'
    ws['D3'] = 'Solde du mois :'
    ws['D3'].font = Font(bold=True)
    ws['E3'] = float(solde_final)
    ws['E3'].number_format = '#,##0.00'

    # --- En-têtes colonnes ---
    headers = ['Date', 'Nature', 'Prestataire/Client', 'Montant encaissé', 'Montant décaissé', 'Commentaire', 'PJ']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
    ws.row_dimensions[5].height = 20

    # --- Données ---
    operations = Operation.objects.filter(
        mois=mois, annee=annee, is_deleted=False
    ).select_related('tiers', 'piece_jointe').order_by('date_operation', 'created_at')

    row = 6
    for op in operations:
        tiers_label = op.tiers.nom if op.tiers else op.tiers_libre
        encaisse = float(op.montant) if op.nature == 'ENTREE' else ''
        decaisse = float(op.montant) if op.nature == 'DEPENSE' else ''
        pj = '📎' if hasattr(op, 'piece_jointe') and op.piece_jointe else ''

        row_data = [
            op.date_operation.strftime('%d/%m/%Y'),
            op.get_nature_display(),
            tiers_label,
            encaisse,
            decaisse,
            op.commentaire,
            pj,
        ]
        bg = BLUE_BG if op.nature == 'ENTREE' else RED_BG
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = thin_border()
            if col in (4, 5) and value != '':
                cell.number_format = '#,##0.00'
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
        row += 1

    # --- Totaux ---
    total_row = row
    ws.cell(total_row, 1, 'TOTAUX').font = Font(bold=True)
    ws.cell(total_row, 4, float(total_entrees)).number_format = '#,##0.00'
    ws.cell(total_row, 5, float(total_depenses)).number_format = '#,##0.00'
    for col in range(1, 8):
        cell = ws.cell(total_row, col)
        cell.fill = PatternFill('solid', fgColor=TOTAL_BG)
        cell.font = Font(bold=True)
        cell.border = thin_border()

    # --- Largeurs colonnes ---
    col_widths = [12, 10, 30, 18, 18, 35, 5]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
